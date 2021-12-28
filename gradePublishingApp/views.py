from django.db.models import Q
from .decarators import allowed_users
from myapp.models import *
from django.shortcuts import render, redirect
import xlwt
from django.http import HttpResponse
from django.contrib.auth import authenticate, logout as django_logout, login as django_login
import bcrypt
import openpyxl


def home(request):
    return render(request, 'index.html')


# def hash_password(plaintext):
#     password = b'plaintext'
#     hashed = bcrypt.hashpw(password, bcrypt.gensalt(14))
#     return hashed

# def compare_password(plain_password, hash):
#     if bcrypt.checkpw(plain_password, hashed):
#         return True
#     else:
#         return False


def dashboard(request):
    print(request.user)
    years = Year.objects.all()
    context = {
        'years': years,
    }
    return render(request, 'dashboard.html', context)


def profile(request):
    return render(request, 'profile.html')

def download_marks(request,course_id):
    response=HttpResponse(content_type='application/ms-excel')
    course=Course.objects.get(id=course_id)
    #decide file name
    response['Content-Disposition'] = f'attachment; filename="{course.code}-marks.xlsx"'

    #creating workbook
    wb = xlwt.Workbook(encoding='utf-8')

    #adding sheet
    ws = wb.add_sheet("sheet1")

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    # headers are bold
    font_style.font.bold = True

    #column header names, you can use your own headers here
    columns = ['mark_id', 'Course', 'reg_number', 'cat marks', 'exam marks']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()
    grades = Marks.objects.filter(course=course_id) #dummy method to fetch data.
    for grade in grades:
        row_num = row_num + 1
        ws.write(row_num, 0, grade.id, font_style)
        ws.write(row_num, 1, grade.course.code, font_style)
        ws.write(row_num, 2, grade.student.reg_number, font_style)
        ws.write(row_num, 3, grade.cat, font_style)
        ws.write(row_num, 4, grade.exam, font_style)
    wb.save(response)
    return response


def levels(request, dep_id):
    return render(request, 'level.html', {'department_id': dep_id})


def grades(request):
    grades=Marks.objects.filter(student=request.user.id)
    context = {
        'grades':grades
    }
    return render(request, 'grades.html', context)


def course_statistics(request,course_id):
    
    if request.method == 'GET':
        grades=Marks.objects.filter(course=course_id)
        context = {
            'grades':grades,
            'course_id':course_id,
            'uploaded': False
        }
        return render(request, 'course_statistics.html', context)
    else: 
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        marks_list = list()
        for row in range(1, len(excel_data)):
            marks_list.append({
                'mark_id': excel_data[row][0],
                'cat': excel_data[row][3],
                'exam': excel_data[row][4],
            })
        for mark in marks_list:
            existing_mark = Marks.objects.filter(id=mark['mark_id'])
            if len(existing_mark) == 0:
                continue
            else: 
                mark_to_update = Marks.objects.get(id=existing_mark[0].id)
                mark_to_update.cat = mark['cat']
                mark_to_update.exam = mark['exam']
                mark_to_update.save()

        grades=Marks.objects.filter(course=course_id)
        context = {
            'grades':grades,
            'course_id':course_id,
            'uploaded': True
        }
        return render(request, 'course_statistics.html', context)



def courses(request, year_id):
    courses = Course.objects.filter(lecture=request.user.id, year=year_id)
    department = Department.objects.get(id=request.user.department.id)
    year = Year.objects.get(id=year_id)
    
    if request.method == 'GET':
        context = {
            'courses': courses,
            'year_id': year_id,
            'uploaded': False
        }
        return render(request, 'courses.html', context)
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        
        student_list = list()
        
        for row in range(1, len(excel_data)):
            student_list.append({
                'name': excel_data[row][1],
                'reg_number': excel_data[row][0],
            })

        for item in student_list:
            existing_student = Student.objects.filter(reg_number=item['reg_number'])
            if len(existing_student) == 0:
                password = item['reg_number']
                student = Student()
                student.name = item['name']
                student.reg_number = item['reg_number']
                student.department = department
                student.year = year
                student.password = password
                student.save()
            else:
                password = item['reg_number']
                student = Student.objects.get(id = existing_student[0].id)
                student.name = item['name']
                student.reg_number = item['reg_number']
                student.department = department
                student.year = year
                student.password = password
                student.save()

        context = {
            'courses': courses,
            'year_id': year_id,
            'uploaded': True
        }
        return render(request, 'courses.html', context)


def course(request, year_id):
    courses = Course.objects.filter(lecture=request.user.id, year=year_id)
    context = {
        'courses': courses,
    }
    return render(request, 'courses.html', context)


def student_login(request):
    if request.method == 'POST':
        student = Student.objects.filter(reg_number=request.POST['reg_number'])
        if len(student) == 0:
            return render(request, 'index.html', {'error': True, 'message': 'Invalid Credentials! ' })
        elif len(student)>0 and student[0].reg_number == request.POST['reg_number'] and student[0].password == request.POST['password']:
            student = Student.objects.get(reg_number=request.POST['reg_number'])
            marks = Marks.objects.filter(student=student.id)
            average = 0
            for mark in marks:
                average += mark.cat + mark.exam
            total_average = average /len(marks)
            context= {
                'marks': marks,
                'student': student,
                'total_average': total_average
            }
            return render(request, 'student-dashboard.html', context)
        else: 
            return render(request, 'index.html', {'error': True, 'message': 'Invalid Credentials! ' })
    else:
        return render(request, 'index.html', {'error': True, 'message': 'Invalid Credentials! '} )


def my_courses(request):
    courses = Course.objects.filter(department = request.user.department.id)
    context = {
        'courses': courses,
    }
    return render(request, 'my_courses.html', context)

def enroll(request,course_id):
    if request.method=='POST':
        course=Course.objects.get(id=course_id)
        if request.POST['enrollment_key']==course.enrollment_key:
            grade=Grade()
            user=User.objects.get(id=request.user.id)
            grade.course=course
            grade.student=user
            grade.save()
            
            context = {
            'course_id': course_id,
            'error': False,
            }
            return render(request, 'enroll.html', context)
            
        else:
            context = {
            'course_id': course_id,
            'error': True,
            }
            return render(request, 'enroll.html', context)
    else:
        context = {
            'course_id': course_id,
            'error': False,
        }
        return render(request, 'enroll.html', context)


def add_course(request):
    years = Year.objects.all()
    groups = Group.objects.filter(Q(name='lecture') | Q(name='hod'))
    leatures = User.objects.filter(group__in=groups)
    
    if request.method == 'POST':
        department = Department.objects.get(id=request.user.department.id)
        lecture= User.objects.get(id=request.POST['lecture'])
        year = Year.objects.get(id=request.POST['year'])
        course = Course()
        course.name = request.POST['name']
        course.code = request.POST['code']
        course.year = year
        course.academic_year = request.POST['academic_year']
        course.department = department
        course.lecture = lecture
        course.save()

        course = Course.objects.filter(name=request.POST['name'], code=request.POST['code'])
        students = Student.objects.filter(year=year.id, department=department.id)
        for student in students:
            mark = Marks()
            mark.course = course[0]
            mark.student = student
            mark.save()
        return render(request, 'add-course.html', {'years': years, 'leatures': leatures, 'twick':True, "message": "Course created successfully"})

    else: 
        return render(request, 'add-course.html',  {'years': years, 'leatures': leatures, 'twick':False, "message": ""})


def my_claims(request):
    claims = Claim.objects.filter(student=request.user.id)
    context = {
        'claims': claims
    }
    return render(request, 'my_claims.html', context)


def page_not_found_view(request, exception):
    return render(request, '404.html', status=404)
